

import openpyxl as opl
from openpyxl.chart.series import Series

#WB = openpyxl.Workbook()
#WB.save('our_first.xlsx')



# wb= openpyxl.load_workbook('our_first.xlsx')
# for sheet in wb:
#     print(sheet.title)


# wb =opl.Workbook()
# ws1 = wb.create_sheet("A_Sheet",0)
# ws2 = wb.create_sheet('B_Sheet',1)
# ws3 = wb.create_sheet('C_Sheet',2)
# ws4 = wb.create_sheet('D_Sheet',3)
# wb.save('our_first.xlsx')
# for sheet in wb:
#     if sheet.title == 'D_Sheet':
#         wb.remove(sheet)
#         break
# wb.save('our_first.xlsx')
# for i in wb:
#     print(i.title)






# wb = opl.load_workbook('our_first.xlsx')
# sheet = wb['Sheet']
# # source = wb['Sheet']
# # wb['A_Sheet'].wb.copy_worksheet(source)
#
#
# ws= wb.worksheets[1]
#
# ws['A1'].value = 'PYTHON'
# ws['B1'].value = 1234
# ws.cell(3,7).value = 'hello'
# ws.cell(3,4).value = 123456
# wb.save('our_first.xlsx')

# for sheet in wb:
#     print(sheet.title)


#wb = opl.load_workbook('our_first.xlsx')
# sheet = wb['Sheet']
#
# for row in sheet.iter_rows(min_row=1,min_col=1,max_row=4,max_col=4,values_only=True):
#     print(row)
#
# for col in sheet.iter_cols(min_row=1,min_col=1,max_row=4,max_col=4,values_only=True):
#     print(col)


#sheet = wb['A_Sheet']
#print(f'max_row : {sheet.max_row} \nmax_col : {sheet.max_column}\n')


#sheet.insert_rows(1)
#sheet.insert_cols(1,3)

#wb.save('our_first.xlsx')
# for row in sheet.iter_rows(values_only=True):
#     print(row)



# appending data into sheet
#wb = opl.load_workbook('our_first.xlsx')
# sheet = wb['B_Sheet']
#
# data = [['apple', 120, 140, 170],
#         ['banana', 80, 90, 95],
#         ['orange',140, 170, 150]]
#
# sheet.append(['name', 2021, 2022, 2023])
# for row in data:
#     sheet.append(row)
# wb.save('our_first.xlsx')
#
# for row in sheet.iter_rows(values_only=True):
#     print(row)



#creating table

from openpyxl.worksheet.table import Table, TableStyleInfo

# wb = opl.load_workbook('our_first.xlsx')
# sheet = wb['C_Sheet']
#
# data = [['apple', 120, 140, 170],
#         ['banana', 80, 90, 95],
#         ['orange',140, 170, 150]]
#
# sheet.append(['name', '2021', '2022', '2023'])
# for row in data:
#     sheet.append(row)
#
# table = Table(displayName = 'Sale_Table', ref = 'A1:D6')
#
# style = TableStyleInfo(name='TableStyleMedium8',showRowStripes=True)
# table.tableStyleInfo = style
# sheet.add_table(table)
# wb.save('our_first.xlsx')



import openpyxl as opl
from openpyxl.chart import *


# Create a new workbook and add data
wb = opl.load_workbook('our_first.xlsx')
sheet = wb['D_Sheet']

# Add some example data to the sheet
data = [
    ["Month", "Sales"],
    ["January", 100],
    ["February", 120],
    ["March", 130],
    ["April", 90],
    ["May", 150]
]

for row in data:
    sheet.append(row)

# Define the data range for the chart
# Data range includes values without headers, so start from row 2 for sales data
data_ref = Reference(sheet, min_col=2, min_row=2, max_row=len(data))

# Define the category range for the chart (month names)
categories_ref = Reference(sheet, min_col=1, min_row=2, max_row=len(data))

# Create a BarChart object
# chart = BarChart()
# chart.title = "Monthly Sales Data"
# chart.x_axis.title = "Months"
# chart.y_axis.title = "Sales"
#
# # Add the data and categories to the chart
# chart.add_data(data_ref, titles_from_data=False)
# chart.set_categories(categories_ref)
#
# # Position the chart in the sheet
# sheet.add_chart(chart, "E5")  # Position the chart starting at cell E5
#
# # Save the workbook
# wb.save('our_first.xlsx')




# Create a Scatterchart object
# chart = ScatterChart()
# chart.title = "Monthly Sales Data"
# chart.x_axis.title = "Months"
# chart.y_axis.title = "Sales"
#
# # Add the data and categories to the chart
# series = Series(data_ref, categories_ref, title_from_data = True)
# chart.series.append(series)
# # Position the chart in the sheet
# sheet.add_chart(chart, "E5")  # Position the chart starting at cell E5
#
# # Save the workbook
# wb.save('our_first.xlsx')


chart = PieChart()
chart.title = "Monthly Sales Data"
chart.x_axis.title = "Months"
chart.y_axis.title = "Sales"

# Add the data and categories to the chart
series = Series(data_ref, categories_ref, title_from_data = True)
chart.series.append(series)
# Position the chart in the sheet
sheet.add_chart(chart, "E5")  # Position the chart starting at cell E5

# Save the workbook
wb.save('our_first.xlsx')